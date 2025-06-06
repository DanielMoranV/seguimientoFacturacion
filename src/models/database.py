import sqlite3
import logging
from pathlib import Path
from typing import Tuple, Any, Dict, List
import pandas as pd
from datetime import datetime
import os
import subprocess

from src.utils.constants import Messages, SQLQueries, ExcelStyles

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # logger.warning(Messages.ERROR_OPENPYXL) # Logger not yet available at module level

class DatabaseManager:
    _instance = None
    
    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
        return cls._instance
    
    def __init__(self, config: Dict, logger: logging.Logger):
        """Iniciar DatabaseManager"""
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.config = config
            self.logger = logger
            self.db_path = self.config['paths']['db_path']
            self.required_columns = self.config['db']['required_columns']
            self.seguimiento_columns = self.config['db']['seguimiento_columns']
            self._setup_database()
            self.logger.info("DatabaseManager inicializado correctamente")
    
    def _setup_database(self):
        """Crear la base de datos SQLite con las tablas necesarias"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS detalle_atenciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                num_doc VARCHAR(10) NOT NULL UNIQUE,
                fec_doc DATE NOT NULL,
                nh_pac VARCHAR(255) NOT NULL,
                nom_pac VARCHAR(255) NOT NULL,
                nom_emp VARCHAR(255) NOT NULL,
                nom_cia VARCHAR(255) NOT NULL,
                ta_doc VARCHAR(1) NOT NULL,
                nom_ser VARCHAR(255) NOT NULL,
                tot_doc DECIMAL(8, 2) NOT NULL,
                num_fac VARCHAR(11) NOT NULL,
                fec_fac DATE NOT NULL,
                num_pag VARCHAR(10) NOT NULL,
                fec_pag DATE NOT NULL,
                usu_sis VARCHAR(255) NOT NULL,
                cod_dx VARCHAR(255) NOT NULL,
                facturador VARCHAR(255) NOT NULL,
                producto VARCHAR(255) NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS seguimiento_facturacion (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                detalle_atencion_id INTEGER NOT NULL,
                estado_aseguradora VARCHAR(255) NULL,
                fecha_envio DATE NULL,
                fecha_recepcion DATE NULL,
                observaciones TEXT NULL,
                acciones VARCHAR(255) NULL,
                FOREIGN KEY (detalle_atencion_id) REFERENCES detalle_atenciones (id)
                    ON DELETE CASCADE
            )
        ''')
        
        conn.commit()
        conn.close()

    def get_stats(self):
        """Obtener estadísticas de la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM detalle_atenciones")
        count = cursor.fetchone()[0]
        conn.close()
        return count

    def clear_database_tables(self) -> Tuple[bool, str]:
        """Limpiar todas las tablas de la base de datos."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM detalle_atenciones")
            cursor.execute("DELETE FROM seguimiento_facturacion") 
            conn.commit()
            conn.close()
            self.logger.info("Todas las tablas de la base de datos han sido limpiadas.")
            return True, "Base de datos limpiada exitosamente."
        except Exception as e:
            self.logger.error(f"Error al limpiar base de datos: {str(e)}")
            return False, f"Error al limpiar base de datos: {str(e)}"

    def export_seguimiento_to_excel(self, export_path: Path) -> Tuple[bool, str]:
        """
        Exportar seguimiento a Excel con formato personalizado
        
        Args:
            export_path: Ruta donde se guardará el archivo Excel
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query(SQLQueries.SELECT_ALL, conn)
            conn.close()

            self.logger.info("Consulta SQL ejecutada correctamente")
            self.logger.info(df.head()) # Log head instead of full df for brevity
            
            total_rows = len(df)
            self.logger.info(f"Total de registros para exportar: {total_rows}")
            
            return self._format_excel(df, export_path)
            
        except Exception as e:
            self.logger.error(f"Error en export_seguimiento_to_excel: {str(e)}")
            return False, Messages.ERROR_EXPORT.format(str(e))
            
    def export_pending_to_excel(self, export_path: Path) -> Tuple[bool, str]:
        """
        Exportar pendientes a Excel con formato personalizado
        (Solo registros sin número de pago y con monto > 0)
        
        Args:
            export_path: Ruta donde se guardará el archivo Excel
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query(SQLQueries.SELECT_PENDING, conn)
            conn.close()

            self.logger.info("Consulta SQL de pendientes ejecutada correctamente")
            self.logger.info(df.head()) # Log head instead of full df for brevity
            
            total_rows = len(df)
            self.logger.info(f"Total de registros pendientes para exportar: {total_rows}")
            
            return self._format_excel(df, export_path)
            
        except Exception as e:
            self.logger.error(f"Error en export_pending_to_excel: {str(e)}")
            return False, Messages.ERROR_EXPORT.format(str(e))
    
    def _format_excel(self, df: pd.DataFrame, export_path: Path) -> Tuple[bool, str]:
        """Aplicar formato al Excel"""
        try:
            # Renombrar columnas usando mapeo de configuración
            column_mapping = self.config['export_columns']
            
            df = df.rename(columns=column_mapping)
            
            # Convertir campos de fecha
            date_columns = self.config['excel']['date_columns']
            for col in date_columns:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            # Convertir campos monetarios
            money_columns = self.config['excel']['money_columns']
            for col in money_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            if not OPENPYXL_AVAILABLE:
                self.logger.warning(Messages.ERROR_OPENPYXL + " No se aplicará formato avanzado.")
                df.to_excel(export_path, index=False, sheet_name=self.config['ui']['export_sheet_name'])
                return True, Messages.SUCCESS_EXPORT.format(str(export_path))

            # Crear el archivo Excel con formato personalizado usando openpyxl
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=self.config['ui']['export_sheet_name'], index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[self.config['ui']['export_sheet_name']]
                
                header_style_config = self.config['excel']['styles']['header']
                header_font = Font(**header_style_config['font'])
                header_fill = PatternFill(start_color=header_style_config['fill']['color'], 
                                          end_color=header_style_config['fill']['color'], 
                                          fill_type='solid') # Corrected fill
                header_alignment = Alignment(**header_style_config['alignment'])
                
                date_style = NamedStyle(name='date_style', number_format=self.config['excel']['styles']['date_format'])
                currency_style = NamedStyle(name='currency_style', number_format=self.config['excel']['styles']['currency_format'])
                
                if 'date_style' not in workbook.named_styles:
                    workbook.add_named_style(date_style)
                if 'currency_style' not in workbook.named_styles:
                    workbook.add_named_style(currency_style)
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                for col_num, column_title in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_num)
                    max_length = max(
                        len(str(column_title)),
                        df[column_title].astype(str).str.len().max() if not df.empty else 0
                    )
                    adjusted_width = min(max(max_length + 2, 10), 50) # Basic auto-adjust
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                    if column_title in date_columns:
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_num)
                            if cell.value is not None: cell.style = date_style
                    elif column_title in money_columns:
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_num)
                            if cell.value is not None: cell.style = currency_style
                
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.freeze_panes = 'A2'
            
            # Abrir el archivo Excel después de exportarlo
            try:
                os.startfile(export_path)
            except Exception as e_open:
                self.logger.warning(f"No se pudo abrir el archivo Excel: {str(e_open)}")
            
            return True, Messages.SUCCESS_EXPORT.format(str(export_path))
            
        except Exception as e:
            self.logger.error(f"Error en _format_excel: {str(e)}")
            return False, Messages.ERROR_EXPORT.format(str(e))

    def validate_excel(self, file_path: str) -> Tuple[bool, pd.DataFrame | None, List[str] | str]:
        """Validar archivo Excel"""
        try:
            df = pd.read_excel(file_path, dtype={'num_doc': str, 'nh_pac': str, 'num_pag': str})
            missing_columns = [col for col in self.required_columns if col not in df.columns]
            if missing_columns:
                return True, df, missing_columns # Return True for valid read, but with missing columns
            return True, df, [] 
        except Exception as e:
            self.logger.error(f"Error al validar Excel {file_path}: {str(e)}")
            return False, None, str(e)

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Limpiar y preparar datos"""
        df_clean = df[self.required_columns].copy()
        df_clean = df_clean.fillna('')
        
        date_columns = ['fec_doc', 'fec_fac', 'fec_pag']
        for col in date_columns:
            try:
                df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce').dt.strftime('%Y-%m-%d')
                df_clean[col] = df_clean[col].fillna('')
            except Exception: # Catch any parsing error
                df_clean[col] = ''
        
        text_fields = ['num_doc', 'nh_pac', 'num_pag']
        for col in text_fields:
            df_clean[col] = df_clean[col].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip().lower() != 'nan' else '')
        
        try:
            df_clean['tot_doc'] = pd.to_numeric(df_clean['tot_doc'], errors='coerce').fillna(0)
        except Exception:
            df_clean['tot_doc'] = 0
        
        return df_clean

    def insert_record(self, cursor: sqlite3.Cursor, row: pd.Series):
        """Insertar registro"""
        query = f'''
            INSERT INTO detalle_atenciones 
            ({', '.join(self.required_columns)})
            VALUES ({', '.join(['?'] * len(self.required_columns))})
        '''
        values = tuple(row[col] for col in self.required_columns)
        cursor.execute(query, values)

    def update_record(self, cursor: sqlite3.Cursor, row: pd.Series, record_id: int):
        """Actualizar registro en la tabla detalle_atenciones"""
        # Excluimos num_doc ya que es el identificador único y no debe cambiar
        update_cols = [col for col in self.required_columns if col != 'num_doc']
        query = f'''
            UPDATE detalle_atenciones 
            SET {', '.join([f'{col}=?' for col in update_cols])}
            WHERE id=?
        '''
        values = tuple(row[col] for col in update_cols) + (record_id,)
        cursor.execute(query, values)

    def process_excel(self, file_path: str, progress_callback: callable) -> Tuple[bool, str]:
        """Procesar archivo Excel con callback de progreso"""
        try:
            is_valid, df, validation_info = self.validate_excel(file_path)
            if not is_valid:
                return False, f"Error al leer archivo: {validation_info}"
            
            if isinstance(validation_info, list) and validation_info: # Check if it's the list of missing columns
                return False, Messages.MISSING_COLUMNS.format(', '.join(validation_info))
            
            if df is None: # Should not happen if is_valid is True, but as a safeguard
                 return False, "Error desconocido al validar el archivo Excel."

            df_clean = self.clean_data(df.copy()) # Use a copy to avoid SettingWithCopyWarning
            total_rows = len(df_clean)
            
            if total_rows == 0:
                return False, Messages.NO_DATA
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            inserted = 0
            updated = 0
            errors = 0
            
            for index, row in df_clean.iterrows():
                try:
                    num_doc = str(row['num_doc']).strip()
                    if not num_doc: # num_doc was already cleaned
                        errors += 1
                        continue
                    
                    cursor.execute(SQLQueries.SELECT_BY_DOC, (num_doc,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        self.update_record(cursor, row, existing[0])
                        updated += 1
                    else:
                        self.insert_record(cursor, row)
                        inserted += 1
                    
                    progress = ((index + 1) / total_rows) * 100
                    progress_callback(progress, f"Procesando: {num_doc}")
                    
                except Exception as e_row:
                    self.logger.error(f"Error procesando fila {index} (num_doc: {num_doc}): {str(e_row)}")
                    errors += 1
                    continue # Continue with the next row
            
            conn.commit()
            conn.close()

            # Actualizar estados de pago después de importar
            payment_success, payment_result = self.update_payment_status()
        
            # Actualizar estados de facturas con monto cero o negativo
            zero_neg_success, zero_neg_result = self.update_zero_negative_status()

            summary = f"Insertados: {inserted}, Actualizados: {updated}, Errores: {errors}"

            if payment_success:
                summary += f"\n{payment_result}"
            else:
                summary += f"\nError al actualizar estados de pago: {payment_result}"
            
            if zero_neg_success:
                summary += f"\n{zero_neg_result}"
            else:
                summary += f"\n{zero_neg_result}"

            return True, summary
            
        except Exception as e_main:
            self.logger.error(f"Error general en process_excel: {str(e_main)}")
            return False, Messages.ERROR_UPDATE.format(str(e_main))

    def update_seguimiento_from_excel(self, file_path: str, progress_callback: callable) -> Tuple[bool, str]:
        """
        Actualizar seguimiento desde archivo Excel
        
        Esta función procesa un archivo Excel con información de seguimiento de facturas y actualiza
        la base de datos. Respeta el estado 'Pagado' de registros existentes y valida datos antes de
        actualizar.
        
        Args:
            file_path (str): Ruta al archivo Excel con datos de seguimiento
            progress_callback (callable): Función para reportar progreso de la operación
            
        Returns:
            Tuple[bool, str]: (Éxito/Fallo, Mensaje descriptivo)
        """
        try:
            self.logger.info(f"Iniciando actualización de seguimiento desde Excel: {file_path}")
            
            # Leer Excel con nombres de columnas amigables para el usuario final
            # Se especifican tipos de datos para columnas críticas para evitar conversiones automáticas incorrectas
            df = pd.read_excel(file_path, dtype={'Número de Documento': str, 'Historia Clínica': str})
            
            # Verificar que todas las columnas requeridas estén presentes en el archivo
            # Usando los nombres amigables definidos en la configuración
            missing_columns = [col for col in self.seguimiento_columns.keys() if col not in df.columns]
            if missing_columns:
                return False, Messages.MISSING_COLUMNS.format(', '.join(missing_columns))
            
            # Seleccionar solo las columnas necesarias y renombrarlas a los nombres de la base de datos
            df_to_process = df[list(self.seguimiento_columns.keys())].copy()
            df_to_process.rename(columns=self.seguimiento_columns, inplace=True)

            # Limpiar datos: convertir NaN a cadenas vacías para evitar problemas con SQLite
            df_clean = df_to_process.fillna('') 
            
            # Convertir y formatear columnas de fecha al formato estándar YYYY-MM-DD
            date_columns_db = ['fecha_envio', 'fecha_recepcion']
            for col in date_columns_db:
                if col in df_clean.columns:
                    try:
                        df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce').dt.strftime('%Y-%m-%d')
                        df_clean[col] = df_clean[col].fillna('') # Asegurar que NAs se conviertan a cadenas vacías
                    except Exception:
                         df_clean[col] = ''
            
            # Verificar que haya datos para procesar
            total_rows = len(df_clean)
            if total_rows == 0:
                return False, Messages.NO_DATA
            
            # Conectar a la base de datos y preparar para procesamiento
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Contadores para el resumen final
            updated_count = 0
            inserted_count = 0
            errors_count = 0
            skipped_paid_count = 0  # Nuevo contador para registros pagados que se omiten
            
            for index, row in df_clean.iterrows():
                try:
                    # Obtener y validar número de documento
                    num_doc = str(row['num_doc']).strip()
                    if not num_doc:
                        errors_count += 1
                        self.logger.warning("Número de documento vacío, omitiendo registro")
                        continue
                    
                    # Buscar el registro correspondiente en detalle_atenciones
                    cursor.execute(SQLQueries.SELECT_BY_DOC, (num_doc,))
                    detalle_record = cursor.fetchone()
                    
                    # Si no existe el registro en detalle_atenciones, no se puede actualizar
                    if not detalle_record:
                        errors_count += 1
                        self.logger.warning(f"No se encontró detalle_atencion para num_doc: {num_doc}")
                        continue
                    
                    detalle_id = detalle_record[0]
                    
                    # Verificar si ya existe un registro de seguimiento para este documento
                    cursor.execute(SQLQueries.SELECT_BY_ID, (detalle_id,))
                    seguimiento_record = cursor.fetchone()
                    
                    # Si existe un registro de seguimiento, verificar si ya está marcado como pagado
                    if seguimiento_record:
                        seguimiento_id = seguimiento_record[0]
                        
                        # Verificar el estado actual del registro
                        cursor.execute(SQLQueries.SELECT_CURRENT_STATUS, (seguimiento_id,))
                        current_status_result = cursor.fetchone()
                        
                        # Si el registro ya está marcado como pagado, no modificarlo
                        if current_status_result and current_status_result[0].strip().lower() == Messages.PAID_STATUS.lower():
                            self.logger.info(f"Omitiendo actualización de registro ya pagado: {num_doc}")
                            skipped_paid_count += 1
                            continue
                    
                    # Preparar datos para SQL, asegurando que todos los campos estén correctamente formateados
                    estado = str(row['estado_aseguradora']).strip()
                    fecha_envio_val = str(row['fecha_envio']).strip() if row['fecha_envio'] else None
                    fecha_recepcion_val = str(row['fecha_recepcion']).strip() if row['fecha_recepcion'] else None
                    observaciones_val = str(row['observaciones']).strip()
                    acciones_val = str(row['acciones']).strip()

                    # Actualizar o insertar el registro según corresponda
                    if seguimiento_record:
                        seguimiento_id = seguimiento_record[0]
                        cursor.execute("""
                            UPDATE seguimiento_facturacion 
                            SET estado_aseguradora = ?, fecha_envio = ?, fecha_recepcion = ?, 
                                observaciones = ?, acciones = ?
                            WHERE id = ?
                        """, (estado, fecha_envio_val, fecha_recepcion_val, observaciones_val, acciones_val, seguimiento_id))
                        updated_count += 1
                    else:
                        cursor.execute("""
                            INSERT INTO seguimiento_facturacion 
                            (detalle_atencion_id, estado_aseguradora, fecha_envio, fecha_recepcion, observaciones, acciones)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (detalle_id, estado, fecha_envio_val, fecha_recepcion_val, observaciones_val, acciones_val))
                        inserted_count += 1
                    
                    # Actualizar barra de progreso
                    progress = ((index + 1) / total_rows) * 100
                    progress_callback(progress, Messages.PROCESSING_DOC.format(num_doc))
                    
                except Exception as e_row_seguimiento:
                    self.logger.error(f"Error procesando seguimiento para num_doc {row.get('num_doc', 'N/A')}: {str(e_row_seguimiento)}")
                    errors_count += 1
                    continue
            
            # Confirmar cambios y cerrar conexión
            conn.commit()
            conn.close()
            
            # Actualizar estados automáticos después de procesar el archivo
            # Primero actualizamos estados de pago
            payment_success, payment_result = self.update_payment_status()
            
            # Luego actualizamos estados de facturas con monto cero o negativo
            zero_neg_success, zero_neg_result = self.update_zero_negative_status()
            
            # Generar resumen de la operación
            summary = Messages.SUCCESS_UPDATE.format(updated_count, inserted_count, errors_count)
            if skipped_paid_count > 0:
                summary += f"\nRegistros ya pagados omitidos: {skipped_paid_count}"
                
            # Añadir resultados de las actualizaciones automáticas
            if payment_success:
                summary += f"\n{payment_result}"
            else:
                summary += f"\nError al actualizar estados de pago: {payment_result}"
                
            if zero_neg_success:
                summary += f"\n{zero_neg_result}"
            else:
                summary += f"\n{zero_neg_result}"
            
            self.logger.info(summary)
            return True, summary
            
        except Exception as e_main_seguimiento:
            self.logger.error(f"Error general en update_seguimiento_from_excel: {str(e_main_seguimiento)}")
            return False, Messages.ERROR_UPDATE.format(str(e_main_seguimiento))

    def update_payment_status(self) -> Tuple[bool, str]:
        """
        Actualizar automáticamente el estado a 'Pagado' en seguimiento_facturacion si hay info de pago en detalle_atenciones.
        
        Esta función busca registros con número de pago válido en detalle_atenciones y actualiza su estado
        a 'Pagado' en la tabla seguimiento_facturacion. No modifica registros que ya estén marcados como pagados.
        
        Returns:
            Tuple[bool, str]: (Éxito/Fallo, Mensaje descriptivo)
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Obtener registros con información de pago
            cursor.execute(SQLQueries.SELECT_PAID)
            paid_records = cursor.fetchall()
            
            updated_count = 0
            inserted_count = 0
            skipped_empty_count = 0  # Contador para registros con num_pag vacío
        
            for record in paid_records:
                detalle_id, num_doc, num_pag, fec_pag = record
                
                # Validar que num_pag no esté vacío (adicional a la consulta SQL)
                if not num_pag or str(num_pag).strip() == '':
                    self.logger.info(f"Omitiendo registro con número de pago vacío: {num_doc}")
                    skipped_empty_count += 1
                    continue
                
                # Asegurar que la fecha de pago sea válida, usar fecha actual si no lo es
                try:
                    valid_fec_pag = pd.to_datetime(fec_pag).strftime('%Y-%m-%d') if fec_pag else datetime.now().strftime('%Y-%m-%d')
                except ValueError:  # Manejar casos donde fec_pag podría ser una cadena de fecha inválida
                    valid_fec_pag = datetime.now().strftime('%Y-%m-%d')

                # Verificar si ya existe un registro de seguimiento para este documento
                cursor.execute(SQLQueries.SELECT_BY_ID, (detalle_id,))
                existing_seguimiento = cursor.fetchone()

                if existing_seguimiento:
                    seguimiento_id = existing_seguimiento[0]
                    
                    # Verificar el estado actual del registro
                    cursor.execute(SQLQueries.SELECT_CURRENT_STATUS, (seguimiento_id,))
                    current_status_result = cursor.fetchone()
                    
                    # No actualizar si ya está marcado como pagado
                    if current_status_result and current_status_result[0].strip().lower() == Messages.PAID_STATUS.lower():
                        continue 

                    # Actualizar el registro existente con estado 'Pagado'
                    cursor.execute("""
                        UPDATE seguimiento_facturacion 
                        SET estado_aseguradora = ?,
                            fecha_recepcion = ?,
                            observaciones = CASE 
                                WHEN observaciones = '' OR observaciones IS NULL THEN ?
                                ELSE observaciones || ' | ' || ?
                            END,
                            acciones = CASE 
                                WHEN acciones = '' OR acciones IS NULL THEN ?
                                ELSE acciones 
                            END
                        WHERE id = ?
                    """, (Messages.PAID_STATUS, valid_fec_pag, 
                          Messages.DEFAULT_OBSERVATION, Messages.DEFAULT_OBSERVATION, 
                          Messages.DEFAULT_ACTION, seguimiento_id))
                    updated_count += 1
                else:
                    # Crear un nuevo registro de seguimiento con estado 'Pagado'
                    cursor.execute("""
                        INSERT INTO seguimiento_facturacion 
                        (detalle_atencion_id, estado_aseguradora, fecha_envio, fecha_recepcion, observaciones, acciones)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (detalle_id, Messages.PAID_STATUS, valid_fec_pag, valid_fec_pag, 
                          Messages.DEFAULT_OBSERVATION, Messages.DEFAULT_ACTION))
                    inserted_count += 1 

            # Confirmar cambios y cerrar conexión
            conn.commit()
            conn.close()
        
            # Generar resumen de la operación
            summary = Messages.SUCCESS_PAYMENT.format(updated_count, inserted_count)
            if skipped_empty_count > 0:
                summary += f" (Omitidos por número de pago vacío: {skipped_empty_count})"
                
            self.logger.info(summary)
            return True, summary
        
        except Exception as e_payment:
            self.logger.error(Messages.ERROR_PAYMENT.format(str(e_payment)))
            return False, Messages.ERROR_PAYMENT.format(str(e_payment))
            
    def update_zero_negative_status(self) -> Tuple[bool, str]:
        """Actualizar automáticamente el estado a 'Cero o Negativo' en seguimiento_facturacion si tot_doc es <= 0."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute(SQLQueries.SELECT_ZERO_NEGATIVE)
            zero_negative_records = cursor.fetchall()
            
            updated_count = 0
            inserted_count = 0
            current_date = datetime.now().strftime('%Y-%m-%d')
        
            for record in zero_negative_records:
                detalle_id, num_doc, tot_doc = record
                
                cursor.execute(SQLQueries.SELECT_BY_ID, (detalle_id,))
                existing_seguimiento = cursor.fetchone()

                if existing_seguimiento:
                    seguimiento_id = existing_seguimiento[0]
                    cursor.execute(SQLQueries.SELECT_CURRENT_STATUS, (seguimiento_id,))
                    current_status_result = cursor.fetchone()
                    
                    # No actualizar si ya tiene estado "Cero o Negativo"
                    if current_status_result and current_status_result[0].strip().lower() == Messages.ZERO_NEGATIVE_STATUS.lower():
                        continue 

                    cursor.execute("""
                        UPDATE seguimiento_facturacion 
                        SET estado_aseguradora = ?,
                            observaciones = CASE 
                                WHEN observaciones = '' OR observaciones IS NULL THEN ?
                                ELSE observaciones || ' | ' || ?
                            END,
                            acciones = CASE 
                                WHEN acciones = '' OR acciones IS NULL THEN ?
                                ELSE acciones 
                            END
                        WHERE id = ?
                    """, (Messages.ZERO_NEGATIVE_STATUS, 
                          Messages.ZERO_NEGATIVE_OBSERVATION, Messages.ZERO_NEGATIVE_OBSERVATION, 
                          Messages.ZERO_NEGATIVE_ACTION, seguimiento_id))
                    updated_count += 1
                else:
                    cursor.execute("""
                        INSERT INTO seguimiento_facturacion 
                        (detalle_atencion_id, estado_aseguradora, fecha_envio, fecha_recepcion, observaciones, acciones)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (detalle_id, Messages.ZERO_NEGATIVE_STATUS, current_date, current_date, 
                          Messages.ZERO_NEGATIVE_OBSERVATION, Messages.ZERO_NEGATIVE_ACTION))
                    inserted_count += 1 

            conn.commit()
            conn.close()
        
            summary = f"Estados 'Cero o Negativo' actualizados: {updated_count}, Nuevos registros: {inserted_count}"
            self.logger.info(summary)
            return True, summary
        
        except Exception as e_zero_neg:
            error_msg = f"Error al actualizar estados 'Cero o Negativo': {str(e_zero_neg)}"
            self.logger.error(error_msg)
            return False, error_msg
