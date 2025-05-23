#!/usr/bin/env python3
"""
Interfaz gráfica moderna para importar datos de Excel a SQLite
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import sqlite3
import threading
import queue
import time
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import os

from config import get_config
from logging_config import setup_logging
from constants import Messages, SQLQueries, ExcelStyles

# Configuración y logging
CONFIG = get_config()
logger = setup_logging()

# Importaciones opcionales para el formato Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning(Messages.ERROR_OPENPYXL)

# Configurar tema y apariencia
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")

class DatabaseManager:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
        return cls._instance
    
    def __init__(self):
        """Iniciar DatabaseManager con configuración centralizada"""
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.db_path = CONFIG['paths']['db_path']
            self.required_columns = CONFIG['db']['required_columns']
            self.seguimiento_columns = CONFIG['db']['seguimiento_columns']
            
            try:
                self.create_database()
                self.update_payment_status()
                logger.info("DatabaseManager inicializado correctamente")
            except Exception as e:
                logger.error(f"Error al inicializar DatabaseManager: {str(e)}")
                raise
    
    def export_seguimiento_to_excel(self, export_path: str) -> Tuple[bool, str]:
        """
        Exportar seguimiento a Excel con formato personalizado
        
        Args:
            export_path: Ruta donde se guardará el archivo Excel
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            logger.info(f"Iniciando exportación a Excel: {export_path}")
            conn = sqlite3.connect(str(self.db_path))
            df = pd.read_sql_query(SQLQueries.SELECT_ALL, conn)
            conn.close()
            logger.info("Consulta SQL ejecutada correctamente")

            # Renombrar columnas usando mapeo de configuración
            column_mapping = CONFIG['export_columns']

            # Renombrar las columnas
            df = df.rename(columns=column_mapping)

            # Convertir campos de fecha a datetime para que Excel los reconozca como fechas
            date_columns = ['Fecha de Documento', 'Fecha de Factura', 'Fecha de Pago', 'Fecha de Envío', 'Fecha de Recepción']
            for col in date_columns:
                if col in df.columns:
                    # Convertir a datetime, manejando valores vacíos y errores
                    df[col] = pd.to_datetime(df[col], errors='coerce')

            # Convertir campos monetarios a numérico
            money_columns = ['Total Documento']
            for col in money_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            # Crear el archivo Excel con formato personalizado
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Seguimiento_Facturacion', index=False)
                
                # Obtener el workbook y worksheet
                workbook = writer.book
                worksheet = writer.sheets['Seguimiento_Facturacion']
                
                # Importar estilos necesarios
                from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
                from openpyxl.utils import get_column_letter
                
                # Crear estilos de cabecera
                header_font = Font(**ExcelStyles.HEADER_FONT)
                header_fill = PatternFill(**ExcelStyles.HEADER_FILL)
                header_alignment = Alignment(**ExcelStyles.HEADER_ALIGNMENT)
                
                # Crear estilos de celda
                date_style = NamedStyle(name='date_style', number_format=ExcelStyles.DATE_FORMAT)
                currency_style = NamedStyle(name='currency_style', number_format=ExcelStyles.CURRENCY_FORMAT)
                
                # Registrar estilos en el workbook
                if 'date_style' not in workbook.named_styles:
                    workbook.add_named_style(date_style)
                if 'currency_style' not in workbook.named_styles:
                    workbook.add_named_style(currency_style)
                
                # Aplicar formato a la cabecera
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Ajustar el ancho de las columnas
                for col_num, column in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_num)
                    max_length = max(
                        len(str(column)),  # Longitud del header
                        df[column].astype(str).str.len().max() if not df.empty else 0  # Longitud del contenido
                    )
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar formato específico a columnas de fecha
                for col_num, column in enumerate(df.columns, 1):
                    column_letter = get_column_letter(col_num)
                    
                    if column in date_columns:
                        for row in range(2, len(df) + 2):  # Empezar desde la fila 2 (después del header)
                            cell = worksheet.cell(row=row, column=col_num)
                            if cell.value is not None:
                                try:
                                    cell.style = date_style
                                except Exception as e:
                                    logger.error(f"Error al aplicar estilo de fecha: {str(e)}")
                    
                    elif column in money_columns:
                        for row in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row, column=col_num)
                            if cell.value is not None:
                                try:
                                    cell.style = currency_style
                                except Exception as e:
                                    logger.error(f"Error al aplicar estilo monetario: {str(e)}")
                
                # Aplicar filtros automáticos
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Congelar la primera fila (cabecera)
                worksheet.freeze_panes = 'A2'

            return True, Messages.SUCCESS_EXPORT.format(export_path)
        except Exception as e:
            logger.error(f"Error en export_seguimiento_to_excel: {str(e)}")
            return False, Messages.ERROR_EXPORT.format(str(e))
    
    def update_seguimiento_from_excel(self, file_path: str, progress_callback: callable) -> Tuple[bool, str]:
        """
        Actualizar seguimiento desde archivo Excel
        
        Args:
            file_path: Ruta del archivo Excel
            progress_callback: Función para actualizar el progreso
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            logger.info(f"Iniciando actualización de seguimiento desde Excel: {file_path}")
            
            # Leer Excel con nombres de columnas amigables
            df = pd.read_excel(file_path, dtype={'Número de Documento': str, 'Historia Clínica': str})
            
            # Mapear nombres amigables a nombres de BD usando configuración
            required_seguimiento_columns = self.seguimiento_columns
            
            # Verificar columnas requeridas
            missing_columns = [col for col in required_seguimiento_columns.keys() if col not in df.columns]
            if missing_columns:
                return False, f"Columnas faltantes: {', '.join(missing_columns)}"
            
            # Limpiar datos
            df_clean = df[list(required_seguimiento_columns.keys())].copy()
            df_clean = df_clean.fillna('')
            
            # Convertir fechas
            date_columns = ['Fecha de Envío', 'Fecha de Recepción']
            for col in date_columns:
                try:
                    df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce').dt.strftime('%Y-%m-%d')
                    df_clean[col] = df_clean[col].fillna('')
                except:
                    df_clean[col] = ''
            
            total_rows = len(df_clean)
            if total_rows == 0:
                return False, "No hay datos válidos para procesar"
            
            # Procesar registros
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            updated = 0
            inserted = 0
            errors = 0
            
            for index, row in df_clean.iterrows():
                try:
                    num_doc = str(row['Número de Documento']).strip()
                    if not num_doc or num_doc == 'nan':
                        errors += 1
                        continue
                    
                    # Buscar el detalle_atencion_id
                    cursor.execute("SELECT id FROM detalle_atenciones WHERE num_doc = ?", (num_doc,))
                    detalle_record = cursor.fetchone()
                    
                    if not detalle_record:
                        errors += 1
                        continue
                    
                    detalle_id = detalle_record[0]
                    
                    # Verificar si existe seguimiento
                    cursor.execute("SELECT id FROM seguimiento_facturacion WHERE detalle_atencion_id = ?", (detalle_id,))
                    seguimiento_record = cursor.fetchone()
                    
                    # Preparar datos
                    estado = str(row['Estado Aseguradora']).strip()
                    fecha_envio = str(row['Fecha de Envío']).strip() if row['Fecha de Envío'] else ''
                    fecha_recepcion = str(row['Fecha de Recepción']).strip() if row['Fecha de Recepción'] else ''
                    observaciones = str(row['Observaciones']).strip()
                    acciones = str(row['Acciones']).strip()
                    
                    if seguimiento_record:
                        # Actualizar seguimiento existente
                        seguimiento_id = seguimiento_record[0]
                        cursor.execute("""
                            UPDATE seguimiento_facturacion 
                            SET estado_aseguradora = ?,
                                fecha_envio = ?,
                                fecha_recepcion = ?,
                                observaciones = ?,
                                acciones = ?
                            WHERE id = ?
                        """, (estado, fecha_envio, fecha_recepcion, observaciones, acciones, seguimiento_id))
                        updated += 1
                    else:
                        # Crear nuevo seguimiento
                        cursor.execute("""
                            INSERT INTO seguimiento_facturacion 
                            (detalle_atencion_id, estado_aseguradora, fecha_envio, fecha_recepcion, observaciones, acciones)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (detalle_id, estado, fecha_envio, fecha_recepcion, observaciones, acciones))
                        inserted += 1
                    
                    # Actualizar progreso
                    progress = ((index + 1) / total_rows) * 100
                    progress_callback(progress, f"Procesando seguimiento: {num_doc}")
                    
                except Exception as e:
                    errors += 1
                    continue
            
            conn.commit()
            conn.close()
            
            summary = f"Seguimientos actualizados: {updated}, Nuevos seguimientos: {inserted}, Errores: {errors}"
            return True, summary
            
        except Exception as e:
            return False, f"Error general: {str(e)}"  
    
    def update_payment_status(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Buscar registros con num_pag (facturas pagadas)
            cursor.execute("""
            SELECT id, num_doc, num_pag, fec_pag 
            FROM detalle_atenciones 
            WHERE num_pag IS NOT NULL 
            AND num_pag != '' 
            AND num_pag != 'nan'
        """)
            
            paid_records = cursor.fetchall()
            updated_count = 0
            inserted_count = 0
        
            for record in paid_records:
                detalle_id, num_doc, num_pag, fec_pag = record
                
                # Verificar si ya existe registro en seguimiento
                cursor.execute("""
                SELECT id FROM seguimiento_facturacion 
                WHERE detalle_atencion_id = ?
                """, (detalle_id,))

                existing_seguimiento = cursor.fetchone()

                if existing_seguimiento:
                    seguimiento_id = existing_seguimiento[0]
    
                    # Verificar si ya está marcado como pagado
                    cursor.execute("""
                        SELECT estado_aseguradora FROM seguimiento_facturacion 
                        WHERE id = ?
                    """, (seguimiento_id,))
                    estado_actual = cursor.fetchone()[0]
    
                    if estado_actual.strip().lower() == 'pagado':
                        continue  # Ya está pagado, no hacer nada

                    # Si no está pagado, actualiza
                    cursor.execute("""
                        UPDATE seguimiento_facturacion 
                        SET estado_aseguradora = 'Pagado',
                            fecha_recepcion = ?,
                            observaciones = CASE 
                                WHEN observaciones = '' OR observaciones IS NULL 
                                THEN 'Estado actualizado automáticamente - Factura pagada'
                                ELSE observaciones || ' | Estado actualizado automáticamente - Factura pagada'
                            END,
                            acciones = CASE 
                                WHEN acciones = '' OR acciones IS NULL 
                                THEN 'Pago procesado'
                                ELSE acciones
                            END
                        WHERE id = ?
                    """, (fec_pag if fec_pag else datetime.now().strftime('%Y-%m-%d'), seguimiento_id))
                    updated_count += 1
                else:
                    # Insertar nuevo registro
                    cursor.execute("""
                        INSERT INTO seguimiento_facturacion (
                            detalle_atencion_id,
                            estado_aseguradora,
                            fecha_envio,
                            fecha_recepcion,
                            observaciones,
                            acciones
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        detalle_id,
                        'Pagado',
                        fec_pag if fec_pag else datetime.now().strftime('%Y-%m-%d'),
                        fec_pag if fec_pag else datetime.now().strftime('%Y-%m-%d'),
                        'Estado actualizado automáticamente - Factura pagada',
                        'Pago procesado'
                    ))
                    inserted_count += 1 

            conn.commit()
            conn.close()
        
            return True, f"Estados actualizados: {updated_count}, Nuevos registros: {inserted_count}"
        
        except Exception as e:
            return False, f"Error al actualizar estados de pago: {str(e)}"
    
    def create_database(self):
        """Crear la base de datos SQLite con las tablas necesarias"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS detalle_atenciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                num_doc VARCHAR(10) NOT NULL UNIQUE,
                fec_doc DATE NOT NULL,
                nh_pac VARCHAR(255) NOT NULL,
                nom_pac VARCHAR(255) NOT NULL,
                nom_emp VARCHAR(255) NOT NULL,
                nom_cia VARCHAR(255) NOT NULL,
                ta_doc VARCHAR(1) NOT NULL,
                nom_ser VARCHAR(255) NOT NULL,
                tot_doc DECIMAL(8, 2) NOT NULL,
                num_fac VARCHAR(11) NOT NULL,
                fec_fac DATE NOT NULL,
                num_pag VARCHAR(10) NOT NULL,
                fec_pag DATE NOT NULL,
                usu_sis VARCHAR(255) NOT NULL,
                cod_dx VARCHAR(255) NOT NULL,
                facturador VARCHAR(255) NOT NULL,
                producto VARCHAR(255) NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS seguimiento_facturacion (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                detalle_atencion_id INTEGER NOT NULL,
                estado_aseguradora VARCHAR(255) NOT NULL,
                fecha_envio DATE NOT NULL,
                fecha_recepcion DATE NOT NULL,
                observaciones TEXT NOT NULL,
                acciones VARCHAR(255) NOT NULL,
                FOREIGN KEY (detalle_atencion_id) REFERENCES detalle_atenciones (id)
                    ON DELETE CASCADE
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def validate_excel(self, file_path):
        """Validar archivo Excel"""
        try:
            # Leer Excel preservando formato de texto en campos específicos
            df = pd.read_excel(file_path, dtype={'num_doc': str, 'nh_pac': str, 'num_pag': str})
            missing_columns = [col for col in self.required_columns if col not in df.columns]
            return True, df, missing_columns
        except Exception as e:
            return False, None, str(e)
    
    def process_excel(self, file_path, progress_callback):
        """Procesar archivo Excel con callback de progreso"""
        try:
            # Validar archivo
            is_valid, df, missing = self.validate_excel(file_path)
            if not is_valid:
                return False, f"Error al leer archivo: {missing}"
            
            if missing:
                return False, f"Columnas faltantes: {', '.join(missing)}"
            
            # Limpiar datos
            df_clean = self.clean_data(df)
            total_rows = len(df_clean)
            
            if total_rows == 0:
                return False, "No hay datos válidos para procesar"
            
            # Procesar registros
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            inserted = 0
            updated = 0
            errors = 0
            
            for index, row in df_clean.iterrows():
                try:
                    num_doc = str(row['num_doc']).strip()
                    if not num_doc or num_doc == 'nan':
                        errors += 1
                        continue
                    
                    # Verificar si existe
                    cursor.execute("SELECT id FROM detalle_atenciones WHERE num_doc = ?", (num_doc,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Actualizar
                        self.update_record(cursor, row, existing[0])
                        updated += 1
                    else:
                        # Insertar
                        self.insert_record(cursor, row)
                        inserted += 1
                    
                    # Actualizar progreso
                    progress = ((index + 1) / total_rows) * 100
                    progress_callback(progress, f"Procesando: {num_doc}")
                    
                except Exception as e:
                    errors += 1
                    continue
            
            conn.commit()
            conn.close()

            # Actualizar estados de pago después de importar
            payment_success, payment_result = self.update_payment_status()

            summary = f"Insertados: {inserted}, Actualizados: {updated}, Errores: {errors}"

            if payment_success:
                summary += f"\n{payment_result}"
            else:
                summary += f"\nError al actualizar estados de pago: {payment_result}"

            return True, summary
            
        except Exception as e:
            return False, f"Error general: {str(e)}"
    
    def clean_data(self, df):
        """Limpiar y preparar datos"""
        df_clean = df[self.required_columns].copy()
        df_clean = df_clean.fillna('')
        
        # Convertir fechas
        date_columns = ['fec_doc', 'fec_fac', 'fec_pag']
        for col in date_columns:
            try:
                df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce').dt.strftime('%Y-%m-%d')
                df_clean[col] = df_clean[col].fillna('')
            except:
                df_clean[col] = ''
        
        # Limpiar campos de texto preservando ceros a la izquierda
        text_fields = ['num_doc', 'nh_pac', 'num_pag']
        for col in text_fields:
            # Convertir a string preservando formato original
            df_clean[col] = df_clean[col].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() != 'nan' else '')
        
        # Convertir tot_doc
        try:
            df_clean['tot_doc'] = pd.to_numeric(df_clean['tot_doc'], errors='coerce').fillna(0)
        except:
            df_clean['tot_doc'] = 0
        
        return df_clean
    
    
    def insert_record(self, cursor, row):
        """Insertar registro"""
        query = '''
            INSERT INTO detalle_atenciones 
            (num_doc, fec_doc, nh_pac, nom_pac, nom_emp, nom_cia, ta_doc, 
             nom_ser, tot_doc, num_fac, fec_fac, num_pag, fec_pag, usu_sis, 
             cod_dx, facturador, producto)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        '''
        cursor.execute(query, tuple(row))
    
    def update_record(self, cursor, row, record_id):
        """Actualizar registro solo si no está marcado como 'Pagado' en seguimiento"""
    
        # Verificar si el estado ya es "Pagado"
        cursor.execute("""
            SELECT estado_aseguradora FROM seguimiento_facturacion 
            WHERE detalle_atencion_id = ?
        """, (record_id,))
        estado = cursor.fetchone()


        if estado and estado[0].strip().lower() == "pagado":
            return  # Ya está pagado, no actualizar nada

        # Si no está pagado, actualizar detalle_atenciones
        query = '''
            UPDATE detalle_atenciones 
            SET fec_doc=?, nh_pac=?, nom_pac=?, nom_emp=?, nom_cia=?, ta_doc=?, 
                nom_ser=?, tot_doc=?, num_fac=?, fec_fac=?, num_pag=?, fec_pag=?, 
                usu_sis=?, cod_dx=?, facturador=?, producto=?
            WHERE id=?
        '''
        values = tuple(row[1:]) + (record_id,)
        cursor.execute(query, values)

    def get_stats(self):
        """Obtener estadísticas de la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM detalle_atenciones")
        count = cursor.fetchone()[0]
        conn.close()
        return count

class ModernExcelImporter(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configuración de ventana
        self.title("Excel to SQLite Importer")
        self.geometry("800x700")
        self.minsize(600, 500)
        
        # Variables
        self.selected_file = None
        self.db_manager = DatabaseManager()
        self.progress_queue = queue.Queue()
        self.selected_seguimiento_file = None
        
        # Configurar interfaz
        self.setup_ui()
        self.update_stats()
        
        # Iniciar verificación de progreso
        self.after(100, self.check_progress_queue)
    
    def setup_ui(self):
        """Configurar interfaz de usuario"""
        # Frame principal
        self.main_frame = ctk.CTkFrame(self, corner_radius=0)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Título
        self.title_label = ctk.CTkLabel(
            self.main_frame, 
            text="Importador de Excel a SQLite",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        self.title_label.pack(pady=(30, 10))

        # Subtítulo
        self.subtitle_label = ctk.CTkLabel(
            self.main_frame,
            text="Selecciona un archivo Excel para importar datos a la base de datos",
            font=ctk.CTkFont(size=14),
            text_color="gray"
        )
        self.subtitle_label.pack(pady=(0, 30))

        # Frame de selección de archivo
        self.file_frame = ctk.CTkFrame(self.main_frame, height=120)
        self.file_frame.pack(fill="x", padx=30, pady=(0, 20))
        self.file_frame.pack_propagate(False)

        # Botón seleccionar archivo
        self.select_button = ctk.CTkButton(
            self.file_frame,
            text="📁 Seleccionar Archivo Excel",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=50,
            command=self.select_file
        )
        self.select_button.pack(pady=20)

        # Label del archivo seleccionado
        self.file_label = ctk.CTkLabel(
            self.file_frame,
            text="Ningún archivo seleccionado",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        self.file_label.pack(pady=(0, 10))

        # Frame de progreso
        self.progress_frame = ctk.CTkFrame(self.main_frame, height=120)
        self.progress_frame.pack(fill="x", padx=30, pady=(0, 20))
        self.progress_frame.pack_propagate(False)

        # Barra de progreso
        self.progress_bar = ctk.CTkProgressBar(self.progress_frame, width=400)
        self.progress_bar.pack(pady=(20, 10))
        self.progress_bar.set(0)

        # Label de progreso
        self.progress_label = ctk.CTkLabel(
            self.progress_frame,
            text="Esperando archivo...",
            font=ctk.CTkFont(size=12)
        )
        self.progress_label.pack(pady=(0, 20))

        # Frame de botones reorganizado y responsivo
        self.button_frame = ctk.CTkFrame(self.main_frame)
        self.button_frame.pack(fill="x", padx=30, pady=(0, 20))

        # Usamos grid para colocar los botones responsivamente
        self.button_frame.columnconfigure((0, 1, 2, 3), weight=1, uniform="a")
        self.button_frame.rowconfigure(0, weight=1)

        # Botón importar
        self.import_button = ctk.CTkButton(
            self.button_frame,
            text="🚀 Importar Datos",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            state="disabled",
            command=self.start_import
        )
        self.import_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        # Botón actualizar seguimiento
        self.update_seguimiento_button = ctk.CTkButton(
            self.button_frame,
            text="🔄 Act. Atenciones",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            fg_color="orange",
            hover_color="darkorange",
            command=self.select_seguimiento_file
        )
        self.update_seguimiento_button.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        # Botón exportar
        self.export_button = ctk.CTkButton(
            self.button_frame,
            text="📥 Exp. Atenciones",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            fg_color="green",
            hover_color="darkgreen",
            command=self.export_seguimiento
        )
        self.export_button.grid(row=0, column=2, padx=10, pady=10, sticky="ew")

        # Botón limpiar base de datos
        self.clear_button = ctk.CTkButton(
            self.button_frame,
            text="🗑️ Limpiar BD",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            fg_color="red",
            hover_color="darkred",
            command=self.clear_database
        )
        self.clear_button.grid(row=0, column=3, padx=10, pady=10, sticky="ew")


        # Frame de estadísticas
        self.stats_frame = ctk.CTkFrame(self.main_frame, height=80)
        self.stats_frame.pack(fill="x", padx=30, pady=(0, 30))
        self.stats_frame.pack_propagate(False)

        # Label de estadísticas
        self.stats_label = ctk.CTkLabel(
            self.stats_frame,
            text="📊 Registros en base de datos: 0",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.stats_label.pack(pady=25)

    def export_seguimiento(self):
        """Exportar seguimiento a Excel"""
        file_path = filedialog.asksaveasfilename(
            title="Guardar archivo Excel",
            defaultextension=".xlsx",
            filetypes=[
                ("Archivos Excel", ".xlsx"),
                ("Todos los archivos", "*.*")
            ]
        )
        if file_path:
            success, message = self.db_manager.export_seguimiento_to_excel(file_path)
            if success:
                messagebox.showinfo("Éxito", message)
            else:
                messagebox.showerror("Error", message)

    
    def select_file(self):
        """Seleccionar archivo Excel"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if file_path:
            self.selected_file = file_path
            filename = os.path.basename(file_path)
            self.file_label.configure(text=f"📄 {filename}")
            self.import_button.configure(state="normal")
            
            # Validar archivo en segundo plano
            threading.Thread(target=self.validate_file_async, daemon=True).start()
 
    def validate_file_async(self):
        """Validar archivo de forma asíncrona"""
        if not self.selected_file:
            return
        
        try:
            is_valid, df, missing = self.db_manager.validate_excel(self.selected_file)
            
            if is_valid and not missing:
                self.progress_queue.put(("validation", "success", f"✅ Archivo válido ({len(df)} filas)"))
            elif missing:
                self.progress_queue.put(("validation", "error", f"❌ Columnas faltantes: {', '.join(missing)}"))
            else:
                self.progress_queue.put(("validation", "error", f"❌ Error en archivo: {missing}"))
        except Exception as e:
            self.progress_queue.put(("validation", "error", f"❌ Error: {str(e)}"))
    
    def start_import(self):
        """Iniciar importación"""
        if not self.selected_file:
            messagebox.showerror("Error", "Por favor selecciona un archivo Excel")
            return
        
        # Deshabilitar botones
        self.import_button.configure(state="disabled", text="⏳ Importando...")
        self.select_button.configure(state="disabled")
        self.clear_button.configure(state="disabled")
        
        # Reiniciar barra de progreso
        self.progress_bar.set(0)
        self.progress_label.configure(text="Iniciando importación...")
        
        # Iniciar importación en hilo separado
        threading.Thread(target=self.import_worker, daemon=True).start()
    
    def import_worker(self):
        """Worker para importación"""
        def progress_callback(progress, message):
            self.progress_queue.put(("progress", progress, message))
        
        try:
            success, result = self.db_manager.process_excel(self.selected_file, progress_callback)
            self.progress_queue.put(("complete", success, result))
        except Exception as e:
            self.progress_queue.put(("complete", False, f"Error inesperado: {str(e)}"))
    
    def check_progress_queue(self):
        """Verificar cola de progreso"""
        try:
            while True:
                msg_type, data1, data2 = self.progress_queue.get_nowait()
                
                if msg_type == "validation":
                    self.progress_label.configure(text=data2)
                    if data1 == "error":
                        self.import_button.configure(state="disabled")
                
                elif msg_type == "progress":
                    progress = data1 / 100.0
                    self.progress_bar.set(progress)
                    self.progress_label.configure(text=data2)
                
                elif msg_type == "complete":
                    success, result = data1, data2
                    self.progress_bar.set(1.0)
                    
                    if success:
                        self.progress_label.configure(text=f"✅ Completado: {result}")
                        messagebox.showinfo("Éxito", f"Importación exitosa!\n\n{result}")
                    else:
                        self.progress_label.configure(text=f"❌ Error: {result}")
                        messagebox.showerror("Error", f"Error en importación:\n\n{result}")
                    
                    # Rehabilitar botones
                    self.import_button.configure(state="normal", text="🚀 Importar Datos")
                    self.select_button.configure(state="normal")
                    self.clear_button.configure(state="normal")
                    
                    # Actualizar estadísticas
                    self.update_stats()
                    
                    # Reiniciar en 3 segundos
                    self.after(3000, self.reset_ui)
                elif msg_type == "seguimiento_complete":
                    success, result = data1, data2
                    self.progress_bar.set(1.0)
                    
                    if success:
                        self.progress_label.configure(text=f"✅ Seguimiento actualizado: {result}")
                        messagebox.showinfo("Éxito", f"Actualización de seguimiento exitosa!\n\n{result}")
                    else:
                        self.progress_label.configure(text=f"❌ Error: {result}")
                        messagebox.showerror("Error", f"Error en actualización:\n\n{result}")

                    # Rehabilitar botones
                    self.update_seguimiento_button.configure(state="normal", text="📝 Actualizar Seguimiento")
                    self.select_button.configure(state="normal")
                    self.import_button.configure(state="normal")
                    self.clear_button.configure(state="normal") 
                    
                    # Actualizar estadísticas
                    self.update_stats()
                    
                    # Reiniciar en 3 segundos
                    self.after(3000, self.reset_ui)
                
        except queue.Empty:
            pass
        
        # Programar siguiente verificación
        self.after(100, self.check_progress_queue)
    
    def reset_ui(self):
        """Reiniciar interfaz"""
        self.progress_bar.set(0)
        self.progress_label.configure(text="Esperando archivo...")
    
    def clear_database(self):
        """Limpiar base de datos"""
        result = messagebox.askyesno(
            "Confirmar",
            "¿Estás seguro de que quieres eliminar todos los registros de la base de datos?\n\nEsta acción no se puede deshacer."
        )
        
        if result:
            try:
                conn = sqlite3.connect(self.db_manager.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM detalle_atenciones")
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Éxito", "Base de datos limpiada exitosamente")
                self.update_stats()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al limpiar base de datos:\n{str(e)}")
    
    def update_stats(self):
        """Actualizar estadísticas"""
        try:
            count = self.db_manager.get_stats()
            self.stats_label.configure(text=f"📊 Registros en base de datos: {count:,}")
        except Exception as e:
            self.stats_label.configure(text="📊 Error al obtener estadísticas")
    def select_seguimiento_file(self):
        """Seleccionar archivo Excel para actualizar seguimiento"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de seguimiento",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if file_path:
            self.selected_seguimiento_file = file_path
            filename = os.path.basename(file_path)
            
            # Deshabilitar botones
            self.update_seguimiento_button.configure(state="disabled", text="⏳ Actualizando...")
            self.select_button.configure(state="disabled")
            self.import_button.configure(state="disabled")
            self.clear_button.configure(state="disabled")
            
            # Reiniciar barra de progreso
            self.progress_bar.set(0)
            self.progress_label.configure(text="Iniciando actualización de seguimiento...")
            
            # Iniciar actualización en hilo separado
            threading.Thread(target=self.update_seguimiento_worker, args=(file_path,), daemon=True).start()

    def update_seguimiento_worker(self, file_path):
        """Worker para actualización de seguimiento"""
        def progress_callback(progress, message):
            self.progress_queue.put(("progress", progress, message))
        
        try:
            success, result = self.db_manager.update_seguimiento_from_excel(file_path, progress_callback)
            self.progress_queue.put(("seguimiento_complete", success, result))
        except Exception as e:
            self.progress_queue.put(("seguimiento_complete", False, f"Error inesperado: {str(e)}"))
def main():
    app = ModernExcelImporter()
    app.mainloop()

if __name__ == "__main__":
    main()